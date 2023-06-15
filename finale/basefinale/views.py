from django.shortcuts import render, redirect
from django import forms
from rest_framework import viewsets
from basefinale.serializers import EntrepriseSerializer, FichierSerializer, EmployeSerializer

from django.contrib.auth.models import User
from basefinale.models import Entreprise, Fichier, Employe

from django.views import View
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.urls import reverse_lazy

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView
from django.contrib.auth.forms import UserCreationForm
from .forms import UserForm
from django.contrib.auth import login

from django.core.mail import send_mail

from django.http import HttpResponse
import openpyxl
import json
import io
import os
from openpyxl import load_workbook
from django.http import JsonResponse


# Create your views here.

class PageLogin(LoginView):
    template_name = 'basefinale/login.html'
    fields = '__all__'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('files')

class RegisterPage(FormView):
    template_name = 'basefinale/register.html'
    form_class = UserCreationForm
    model = User
    #fields = ['username', 'password1', 'password2', 'email']
    redirect_authenticated_user = True
    success_url = reverse_lazy('files')

    def form_valid(self, form):
        user = form.save()
        if user is not None:
            login(self.request, user)
        return super(RegisterPage, self).form_valid(form)
        ##S'assurer ici que l'utilisateur est authentifié

    def get(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            return redirect('files')
        return super(RegisterPage, self).get(*args, **kwargs)

    #def get(self, request):
    #    form = UserForm()
    #    return render(request, 'basefinale/register.html', {'form': form})

    #def post(self, request):
    #    form = UserCreationForm(request.POST)
    #    if form.is_valid():
    #        form.save()
    #        return redirect('files')
    #    return render(request, 'basefinale/register.html', {'form': form})


    #def form_valid(self, form):
    #    user = form.save()
    #    if user is not None:
    #        login(self.request, user)
    #    return super(RegisterPage, self).form_valid(form)

class FichierViewSet(viewsets.ModelViewSet):

    queryset = Fichier.objects.all()
    serializer_class = FichierSerializer
    filterset_fields = ['favourite', 'company']
    search_fields = ['title']


class EntrepriseViewSet(viewsets.ModelViewSet):

    queryset = Entreprise.objects.all()
    serializer_class = EntrepriseSerializer
    search_fields = ['name']

class EmployeViewSet(viewsets.ModelViewSet):

    queryset = Employe.objects.all()
    serializer_class = EmployeSerializer
    search_fields = ['identifier']


class FileList(LoginRequiredMixin, ListView):
    model = Fichier
    context_object_name = 'files'
    template_name = 'basefinale/fichier_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #context['tasks'] = context['tasks'].filter(user=self.request.user)
        #context['count'] = context['tasks'].filter(complete=False).count()
        
        ##Gestion de la recherche dans la page FileList
        ##title__startswith=search_input si on veut les fichiers qui commencent par la valeur entrée
        ##title__contains=search_input si on veut les fichiers contenant la valeur entrée

        search_input = self.request.GET.get('search-area') or ''
        if search_input:
            context['files'] = context['files'].filter(title__contains=search_input)
        
        ##Conserver la valeur de la dernière recherche dans la barre de recherche
        context['search_input'] = search_input
        return context


class FileDetail(LoginRequiredMixin, DetailView):
    model = Fichier
    context_object_name = 'file'

    #def get_context_data(self, **kwargs):
    #    context = super().get_context_data(**kwargs)

        # Charger le fichier Excel et extraire les données
    #    file = self.object.content.path
    #    wb = load_workbook(filename=file, read_only=True)
    #    ws = wb.active
    #    rows = []
    #    for row in ws.iter_rows():
    #        row_data = [cell.value for cell in row]
    #        rows.append(row_data)
    #    wb.close()

        # Sérialiser les données en JSON et les ajouter au contexte
    #    data = json.dumps(rows)
    #    context['data'] = data

    #    return context

class FileContent(LoginRequiredMixin, DetailView):
    model = Fichier
    context_object_name = 'file'
    #template_name = 'basefinale/contenu.html'

    def get(self, request, *args, **kwargs):
        file = self.get_object()

        #filename = "mon_fichier.xlsx"
        filename = file.content.path
        extension = os.path.splitext(filename)[1]
            
        if extension == ".xls" or extension == ".xlsx":
            print("Le fichier est un fichier Excel")
            wb = load_workbook(filename=file.content.path, read_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows():
                row_data = [cell.value for cell in row]
                if any(row_data):  # only add non-empty rows
                    rows.append(row_data)
            wb.close()
            #data = json.dumps(rows)
            data = rows
            context = {'data': data}
            return render(request, 'basefinale/contenu.html', context)
        else:
            print("Le fichier n'est pas un fichier Excel")
            image = file.content
            context = False
            #return render(request, 'basefinale\templates\basefinale\fichier_detail.html')

       

class FileAdd(LoginRequiredMixin, CreateView):
    model = Fichier
    fields = ['title', 'favourite', 'content', 'company']
    success_url = reverse_lazy('files') #renvoie à la liste des fichiers après l'upload d'un fichier
    template_name = 'basefinale/upload_fichier.html'

    def form_valid(self, form):
        form.instance.user = self.request.user
        result = super(FileAdd, self).form_valid(form)

        # Récupérer le fichier uploadé par l'utilisateur
    #    fichier = form.cleaned_data.get('content')
        
        # Charger le fichier avec openpyxl
    #    wb = load_workbook(fichier, read_only=True)
        
        # Sélectionner la feuille de travail à extraire
    #    ws = wb.active
        
        # Extraire les données de la feuille de travail
    #    data = []
    #    for row in ws.iter_rows(values_only=True):
    #        data.append(row)
        
        # Fermer le fichier Excel
    #    wb.close()
        
        # Retourner les données sous forme de JSON
    #    return JsonResponse({'data': data})

        # Envoi d'un mail à tous les utilisateurs
        #subject = 'Nouveau fichier ajouté sur CONFIDENZ'
        #message = "L'untilisateur {} a ajouté un nouveau fichier : {} . Pensez à le consulter".format(self.request.user.username, form.instance.title)
        #from_email = 'erwinbidzana@gmail.com'
        #recipient_list = [u.email for u in User.objects.all()]
        #send_mail(subject, message, from_email, recipient_list)

        return result

    #def upload_file(request):
    #    if request.method == 'POST':
    #        file = request.FILES['file']
    #        wb = openpyxl.load_workbook(file)
    #        ws = wb.active
    #        data = []
    #        for row in ws.iter_rows():
    #            row_data = []
    #            for cell in row:
    #                row_data.append(str(cell.value))
    #            data.append('\t'.join(data))
    #        file_data = '\n'.join(data)
    #        response = HttpResponse(content_type='text/plain')
    #        response['Content-Disposition'] = 'attachment; filename="data.txt"'
    #        response.write(file_data)
    #        return response
    #    return render(request, 'upload_fichier.html')

class FileDelete(LoginRequiredMixin, DeleteView):
    model = Fichier
    context_object_name = 'fichier'
    success_url = reverse_lazy('files')
    template_name = "basefinale/supprimer_fichier.html"

class FileUpdate(LoginRequiredMixin, UpdateView):
    model = Fichier
    fields = ['title', 'favourite', 'content', 'company']
    success_url = reverse_lazy('files')
    template_name = 'basefinale/modifier_fichier.html'

class CompanyAdd(LoginRequiredMixin, CreateView):
    model = Entreprise
    fields = '__all__'
    success_url = reverse_lazy('files')
    template_name = 'basefinale/ajout_entreprise.html'

class EmployeAdd(LoginRequiredMixin, CreateView):
    model = Employe
    fields = ['company', 'identifier']
    success_url = reverse_lazy('files')
    template_name = 'basefinale/ajout_employe.html'

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super(EmployeAdd, self).form_valid(form)

